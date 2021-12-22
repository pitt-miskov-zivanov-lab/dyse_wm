import re
import random
import openpyxl
import math
import pandas as pd
import numpy as np
from io import StringIO
import logging
from Translation.model import get_model, model_to_dict
from collections import deque
import ast

# define regex for regulator update functions
_VALID_CHARS = r'a-zA-Z0-9\_'

####################################################################
############ 			Simulator object  		 		############
####################################################################

class Simulator(object):
	""" Define a simulation object, containing simulation parameters and elements.

		The Simulator object contains dictionaries of:
			- Element objects
			- initial values
			- toggle times and values
			- update groups
		Any element-specific model characteristics are properties of the Element objects
	"""

	# TODO: use __slots__ for memory optimization
	# TODO: use classmethods to read in from a model file or from a dataframe
	def __init__(self, model_file):
		""" Initialize the model object using model information from the input file
			Inputs:
				model_file : an excel spreadsheet containing required columns:
					Variable Name, Positive, Negative, Initial
		"""

		global _VALID_CHARS

		# Model defaults
		# TODO: make inputs to the __init__ function?
		# TODO: check issues with mutable types
		# Default number of levels
		DEF_LEVELS = 3
		# Default increment function for updating element value
		DEF_INCREMENT = 1
		# Default delay noise value
		DEF_TAU_NOISE = 1
		# Default delta for the random-delay sync update scheme
		DEF_DELAY_DELTA = 2
		# Default to decrease when regulation scores are equal
		DEF_BALANCING = ['decrease', '0']
		# Default to zero delay spontaneous activation or inhibition
		# for elements with no positive or negative regulators, respectively
		DEF_SPONT = 0
		# Default update probability
		DEF_UPDATE_PROBABILITY = 0.5
		# Default update rank
		DEF_UPDATE_RANK = 0

		# Initialize model properties
		# List of all elements in the model
		self.__getElement = dict()
		# List of elements with regulators
		self.__updateList = list()
		# Initial values
		self.__initial = dict()
		# Random initial values
		self.__randomInitial = dict()
		# Knockout initial values
		self.__knockout = dict()
		# List of elements that toggle
		self.__switchStep = dict()
		# Toggle values for toggle elements
		self.__switchValue = dict()
		# Element group values for group updating
		self.__groupUpdate = dict()
		# List of groups being used in the model
		self.__groups = list()
		# Element priority values for update rate
		# Higher value means it should be updated more often
		# Used in RA simulation only
		self.__rateUpdate = dict()
		# List of element names to use for rate update,
		# where the number of times the name appears in this list
		# is proportional to the update rate
		self.__rateUpdateList = []
		# If this remains 0 we are not using update rate,
		# if > 0 we are using update rate
		self.__totalPriority = 0
		# Element rank indicates which should be updated first
		# High rank means updated first
		# No rank means updated last
		# Used in Synchronous simulation only
		self.__rankUpdate = dict()
		# Store element probabilities for ra multi
		self.__probUpdate = dict()

		# Load the input file containing elements and regulators
		model_sheets = pd.ExcelFile(model_file)

		# get the model from the first sheet, will check the other sheets for truth tables later
		df_model = model_sheets.parse(
			0, na_values='NaN', keep_default_na=False, index_col=None)

		# check model format
		if 'element attributes' in [x.lower() for x in df_model.columns]:
			df_model = df_model.rename(
					columns=df_model.iloc[1]
					).drop([0,1]).set_index('#')

		# get other sheets, will parse truth tables later
		if len(model_sheets.sheet_names) > 1:
			df_other_sheets = {
					sheet : model_sheets.parse(
							sheet,na_values='NaN',keep_default_na=False
							)
					for sheet in model_sheets.sheet_names[1:]
					}
		else:
			df_other_sheets = ''

		# required columns
		input_col_X = [
				x.strip() for x in df_model.columns
				if ('variable' in x.lower())
				]
		input_col_A = [
				x.strip() for x in df_model.columns
				if ('positive' in x.lower())
				]
		input_col_I = [
				x.strip() for x in df_model.columns
				if ('negative' in x.lower())
				]
		input_col_initial = [
				x.strip() for x in df_model.columns
				if ('initial' in x.lower()
				or 'scenario' in x.lower())
				]

		unique_init = set(input_col_initial)
		for u_init in unique_init:
			if input_col_initial.count(u_init) > 1:
				raise ValueError('Duplicate scenario column names')

		# check for all required columns or duplicate colummns
		if (len(input_col_X) == 0
				or len(input_col_A) == 0
				or len(input_col_I) == 0
				or len(input_col_initial) == 0
				):
			raise ValueError(
					'Missing one or more required columns in input file: '
					'Variable, Positive, Negative, Scenario'
					)
		elif (len(input_col_X) > 1
				or len(input_col_A) > 1
				or len(input_col_I) > 1
				):
			raise ValueError(
					'Duplicate column of: '
					'Variable, Positive, Negative'
					)

		# optional columns
		input_col_maxstate = [
				x for x in df_model.columns
				if (('states' in x.lower()) or ('levels' in x.lower()))
				]
		input_col_increment = [
				x for x in df_model.columns
				if ('increment' in x.lower())
				]
		input_col_delays = [
				x for x in df_model.columns
				if (('timing' in x.lower()) or ('delay' in x.lower()))
				]
		input_col_spont = [
				x for x in df_model.columns
				if ('spontaneous' in x.lower())
				]
		input_col_balance = [
				x for x in df_model.columns
				if ('balancing' in x.lower())
				]
		input_col_update_group = [
				x for x in df_model.columns
				if ('update group' in x.lower())
				]
		input_col_update_rate = [
				x for x in df_model.columns
				if ('update rate' in x.lower())
				]
		input_col_update_rank = [
				x for x in df_model.columns
				if ('update rank' in x.lower())
				]
		input_col_update_prob = [
				x for x in df_model.columns
				if ('update probability' in x.lower())
				]
		input_col_in_out = [
				x for x in df_model.columns
				if ('optimization input' in x.lower())
				]
		input_col_in_out_value = [
				x for x in df_model.columns
				if ('optimization fixed' in x.lower())
				]
		input_col_obj_weight = [
				x for x in df_model.columns
				if ('optimization objective' in x.lower())
				]
		input_col_tau_noise = [
			x for x in df_model.columns if ('noise' in x.lower())]
		input_col_delta = [
			x for x in df_model.columns if ('delta' in x.lower())]
		# experimental or historical data
		input_col_data = [
			x for x in df_model.columns if ('data' in x.lower())]

		if input_col_data != '':
			self.__exp_data_list = [{} for i in range(len(input_col_data))]


		# Initialize lists for storing special initial values
		# for each scenario (randomizing, knockout)
		for scenario,init_col in enumerate(input_col_initial):
				self.__randomInitial[scenario] = []
				self.__knockout[scenario] = []

		# Parse each row of the input model file
		# TODO: use to_dict (reference model.py) to iterate faster
		for curr_row in df_model.index:
			# Each row in the spreadsheet corresponds to one element

			# Get names of the element (X), activators (A), and inhibitors (I)
			X = df_model.loc[curr_row, input_col_X[0]].strip()
			# TODO: remove below check when above parsing is replaced with Translation.get_model
			if set('~!#$%^&*()-\ []\{\}+=\'\"\|/?<>,.:;').intersection(X):
				raise ValueError('Invalid characters in variable name: {}'
						'\n\tUse only letters, numbers, and underscores'.format(X))
			A = df_model.loc[curr_row, input_col_A[0]].strip()
			I = df_model.loc[curr_row, input_col_I[0]].strip()

			# Check for duplicate variables
			if X in self.__getElement:
				raise ValueError('Duplicate variable name: ' + str(X))

			# Get the number of levels for this element
			levels = DEF_LEVELS
			if len(input_col_maxstate) > 0:
				if df_model.loc[curr_row, input_col_maxstate[0]] is not '':
					levels = int(df_model.loc[curr_row, input_col_maxstate[0]])

			# Add this element to the list of elements that will be updated if it has regulators
			if A != '' or I != '':
				self.__updateList += [X]
			else:
				# Look for truth tables in other sheets
				if df_other_sheets != '':
					for key, sheet in df_other_sheets.items():
						if X in sheet.keys():
							# TODO: this function should probably belong to
							# the Element object instead,
							# need to enable creating the node object first
							# and then setting properties
							A = self.parse_truth_table(sheet,levels)
							self.__updateList += [X]

			# Get the increment function for this element
			increment = DEF_INCREMENT
			if len(input_col_increment) > 0:
				if df_model.loc[curr_row, input_col_increment[0]] is not '':
					increment = float(df_model.loc[curr_row, input_col_increment[0]])

			# Get the length of the allowed noise for random delays
			noise = DEF_TAU_NOISE
			if len(input_col_tau_noise) > 0:  # Check for the delay noise column
				if df_model.loc[curr_row, input_col_tau_noise[0]] is not '':
					noise = int(df_model.loc[curr_row, input_col_tau_noise[0]])

			# Get delta for the random-delay sync update scheme
			delta = DEF_DELAY_DELTA
			if len(input_col_delta) > 0:  # Check for the delay noise column
				if df_model.loc[curr_row, input_col_delta[0]] is not '':
					delta = int(df_model.loc[curr_row, input_col_delta[0]])

			# Values for steady state optimization analysis
			opt_input = False
			opt_input_value = ''
			opt_output = False
			opt_output_value = ''
			opt_obj_weight = ''
			if len(input_col_in_out) > 0:
				if df_model.loc[curr_row, input_col_in_out[0]] is not '':
					if df_model.loc[curr_row, input_col_in_out[0]] in ['I', 'i', 'Input', 'input']:
						opt_input = True
					elif df_model.loc[curr_row, input_col_in_out[0]] in ['O', 'o', 'Output', 'output']:
						opt_output = True

			if len(input_col_in_out_value) > 0:
				if df_model.loc[curr_row, input_col_in_out_value[0]] is not '':
					fixed_value = df_model.loc[curr_row, input_col_in_out_value[0]]
					if opt_input:
						opt_input_value = fixed_value
					elif opt_output:
						opt_output_value = fixed_value

			if len(input_col_obj_weight) > 0:
				if df_model.loc[curr_row, input_col_obj_weight[0]] is not '':
					opt_obj_weight = df_model.loc[curr_row, input_col_obj_weight[0]]

			# Get initial value for this element and any toggling
			self.__switchValue[X] = dict()
			self.__switchStep[X] = dict()
			self.__initial[X] = dict()

			for scenario, init_col in enumerate(input_col_initial):
				initial_value_input = str(df_model.loc[curr_row, init_col])

				initial_value_split = initial_value_input.split(',')
				if len(initial_value_split) > 1:
					# Get toggle time and value
					initial_value_input = initial_value_split[0]
					self.__switchValue[X][scenario] = [
							int(t.split('[')[0])
							for t in initial_value_split[1:]
							]
					self.__switchStep[X][scenario] = [
							int(t.split('[')[1][:-1])
							for t in initial_value_split[1:]
							]
				else:
					self.__switchValue[X][scenario] = []
					self.__switchStep[X][scenario] = []

				if str(initial_value_input).lower().strip() in ['r', 'random']:
					# add to list of variables to be randomized
					self.__randomInitial[scenario] += [X]
					init_val = random.randrange(levels)
				elif str(initial_value_input).lower().strip() in ['l', 'low']:
					init_val = 0
				elif str(initial_value_input).lower().strip() in [
						'm','med','medium','middle','moderate'
						]:
					init_val = int(levels/2)
				elif str(initial_value_input).lower().strip() in ['h', 'high']:
					init_val = levels-1
				elif str(initial_value_input).lower().strip() in ['x']:
					self.__knockout[scenario] += [X]
					init_val = 0
				elif initial_value_input is not '':
					init_val = int(initial_value_input)
				else:
					raise ValueError(
							'Missing scenario {} initial value '
							'for element {}'.format(scenario, X)
							)

				if init_val < levels:
					# init_val will be used as an index
					self.__initial[X][scenario] = init_val
				else:
					raise ValueError('Invalid initial value for element {}'.format(X))

			# Get element experimental or historical data
			if input_col_data != '':
				for scenario, data_col in enumerate(input_col_data):
					scenario_data = str(df_model.loc[curr_row, data_col])
					if scenario_data != '':
						scenario_data_split = scenario_data.split('-')
						# Converting string to list
						x_axis = ast.literal_eval(scenario_data_split[0])
						y_axis = ast.literal_eval(scenario_data_split[1])
						self.__exp_data_list[scenario][X] = [x_axis, y_axis]

			# Get element state transition delays;
			# different delays can be defined for each level transition
			# (rising or falling edge)

			delays = [0 for x in range(2*(levels-1))]
			if len(input_col_delays) > 0:
				if df_model.loc[curr_row, input_col_delays[0]] is not '':
					delays = [int(x.strip()) for x in str(
						df_model.loc[curr_row, input_col_delays[0]]).split(',')]
			# check that the number of delays matches the number of states
			if len(delays) != 2*(levels-1):
				if len(delays) == 1:
					# use the same delay for all transitions
					delays = [int(delays[0]) for x in range(2*(levels-1))]
				else:
					# incorrect delay format in the input file
					raise ValueError(
							'For element delays, need 2*(levels - 1) '
							'delays in the format: \n '
							'delay01,delay12,delay21,delay10 \n '
							'For example, for a 3 state model '
							'(levels 0, 1, 2): 1,2,2,1 \n '
							'OR, only one delay value, which will be used '
							'for all level transitions \n '
							'OR, leave blank for no delay'
							)

			# Get balancing behavior
			balancing = DEF_BALANCING
			if len(input_col_balance) > 0:
				if df_model.loc[curr_row,input_col_balance[0]] is not '':
					balancing_input = [
							x.strip() for x in str(
									df_model.loc[curr_row,input_col_balance[0]]
									).split(',')
							]
					if (len(balancing_input) == 2
							and balancing_input[0]
							in ['increase','decrease','positive','negative']
							):
						balancing = balancing_input
					elif balancing_input[0] in ['-', 'none', 'None']:
						balancing = ''
					else:
						raise ValueError('Balancing behavior in the input file must be in the format: '
								'[increase/decrease],[delay]'
								'\n\tFor example: increase,0'
								'\n\tFor no balancing, input None')

			# Get spontaneous activation/inhibition behavior and delays
			spont_delay = DEF_SPONT
			if len(input_col_spont) > 0:
				if df_model.loc[curr_row, input_col_spont[0]] is not '':
					if df_model.loc[curr_row, input_col_spont[0]] not in ['-', 'none', 'None']:
						spont_delay = int(df_model.loc[curr_row, input_col_spont[0]])
					else:
						# spontaneous input is None, set as blank to indicate no spontaneous behavior
						# distinguishing this case from the default value DEF_SPONT, which is spontaneous behavior with 0 delay
						spont_delay = ''

			# only get update group, rate, rank, prob if this element has regulators
			if X in self.__updateList:

				# Set this element's update group for simulations
				if len(input_col_update_group) > 0:
					if df_model.loc[curr_row, input_col_update_group[0]] is not '':
						group = df_model.loc[curr_row, input_col_update_group[0]]
						self.__groupUpdate[X] = group
						if group not in self.__groups:
							self.__groups += [group]

				# Set this element's update rate for simulations
				if len(input_col_update_rate) > 0:
					if df_model.loc[curr_row, input_col_update_rate[0]] is not '':
						rate = df_model.loc[curr_row, input_col_update_rate[0]]
					else:
						# Default rate
						rate = 1

					self.__rateUpdate[X] = rate

					# Include instances of element X in the array
					# proportional to update rate
					self.__rateUpdateList += [X]*rate

					# Increase total priority by the rate
					self.__totalPriority += rate

				# Set this element's update rank for round-based simulations
				if len(input_col_update_rank) > 0:
					if df_model.loc[curr_row, input_col_update_rank[0]] is not '':
						rank = df_model.loc[curr_row, input_col_update_rank[0]]
					else:
						rank = DEF_UPDATE_RANK
					# If this rank already exists in our dictionary, just append the current element into this ranking
					if rank in self.__rankUpdate:
						self.__rankUpdate[rank] += [X]
					# Else create a new entry in the dictionary to store everything in this rank
					else:
						self.__rankUpdate[rank] = [X]

				# Set this element's update probability for random asynchronous multi-step simulations
				prob = DEF_UPDATE_PROBABILITY
				if len(input_col_update_prob) > 0:
					if df_model.loc[curr_row, input_col_update_prob[0]] is not '':
						prob = df_model.loc[curr_row, input_col_update_prob[0]]
					self.__probUpdate[X] = prob

			# Create an element/node object for this element
			ele = Element(
					X, A, I, self.__initial[X][0],
					levels, delays, balancing,
					spont_delay, noise, delta,
					opt_input, opt_output,
					opt_input_value, opt_output_value, opt_obj_weight,
					increment
					)

			# Include this element in the model's dictionary of elements
			self.__getElement[X] = ele

			# Go to the next row in the input file
			curr_row += 1

		# check that all regulators are in the model's element list
		# NOTE: 'for name in self.__getElement' iterates over the keys
		# faster than calling __getElement.keys()
		for key in self.__getElement:
			for reg in self.__getElement[key].get_name_list():
				if reg not in self.__getElement:
					if type(reg) is not str:
						raise ValueError(
								'Invalid regulator value type '
								'for element {}'.format(key)
								)
					else:
						raise ValueError(
								'No Variable Name for element {}'.format(reg)
								)

	## Simulator get/set functions

	def get_elements(self):
		return self.__getElement

	def get_initial(self):
		return self.__initial

	def get_exp_data(self):
		return self.__exp_data_list

	def set_initial(self, scenario=0):
		""" Set the current value of each element (node) in the model
			to its initial value
		"""
		for key, element in self.__getElement.items():
			val = self.__initial[key].get(scenario)
			if val is not None:
				element.set_value_index(val)
			else:
				raise ValueError('Scenario {} does not exist in model '
					'(note scenario is zero-indexed, 0 is the first scenario)'.format(scenario))

	def set_random_initial(self, scenario=0):
		""" Randomize the initial values of elements for the synchronous/simultaneous
			scheme to low, medium, or high based on the number of levels
		"""
		# Even though the initial state is randomized when the Simulator is initialized,
		# this function is needed so you can run multiple simulations
		# with random initial values using the same Simulator object
		for name in self.__randomInitial[scenario]:
			init_val_index = random.randrange(self.__getElement[name].get_levels())
			self.__getElement[name].set_value_index(init_val_index)

	def knockout(self, scenario=0):
		"""Remove all instances of knockout elements from the model
		"""
		for name in self.__knockout[scenario]:
			# knock out this element TODO: need to remove from all interactions
			raise ValueError('Knockout function is not yet supported: {}'.format(name))

	## Simulator simulation functions

	def run_simulation(self,
					simtype,
					runs,
					simStep,
					outName,
					scenario=0,
					outMode=1,
					normalize=False,
					randomizeEachRun=False,
					eventTraces=None,
					progressReport=False,
					):
		""" Run a simulation!
			Inputs
				simtype : simulation scheme
					TODO: update to new terminology
					'ra' : random asynchronous
					'round' : round based
					'sync' : synchronous / simultaneous
					'ra_multi' : random asynchronous multi-step
					'sync_multi' : synchronous multi-step
					'rand_sync':random delay synchronous [uniform distribution]
					'rand_syc_gauss': random delay synchronous [Gaussian distribution]
					'fixed_updates': update order according to event trace input
				runs : number of simulation runs
				simStep : number of simulation steps
				outName : name of output file
				scenario : index of initial value column (zero-indexed)
				outMode : specify output mode (1: all runs and summary, 2: transpose format, 3: summary only)
				normalize : whether to output values normalized to the range [0,1]
				randomizeEachRun : for ra and ra_multi specify whether or not to randomize initial values at beginning of each new run
				eventTraces: name of the event traces file for the fixed_updates scheme
				progressReport: boolean variable specifying whether to report the percentage of completed simulation runs
		"""
		# # Timing code below, to separate simulation time from parsing
		# import time
		# starttime = time.time()

		if simtype not in ['ra', 'round', 'sync', 'ra_multi', 'sync_multi', 'rand_sync', 'rand_sync_gauss','fixed_updates']:
			raise ValueError(
				'Invalid simulation scheme, must be ra, round, sync, ra_multi, sync_multi, rand_sync, rand_sync_guess or fixed_updates')

		if simtype == 'fixed_updates':
			# Get event traces from file input to set element update order,
			# number of steps and runs
			# TODO: move file parsing out of simulator, pass in event trace object
			with open(eventTraces) as event_traces_file:
				content = event_traces_file.readlines()
			content = [x.strip() for x in content]

			numLines = len(content)

			updates = dict()

			for content_index in range(numLines):
				# check each line for the Run #,
				# save the last run # as the total number of runs
				if re.match(r"Run #[0-9]+",content[content_index]):
					runs = int(re.findall(r"Run #([0-9]+)",content[content_index])[0])
				else:
					updates[runs] = content[content_index].split()

			if len(updates[runs]) != simStep:
				logging.warn('Steps input does not match event trace file.'
						'\n\tUsing number of steps in event traces: {}'.format(updates[runs]))
				simStep = len(updates[runs])
			# runs are 0-indexed
			# increment the value read from the file to get the total number of runs
			runs+=1

		with open(outName, 'w') as output_file:

			# Set elements to initial values
			self.set_initial(scenario)

			# freq_sum will keep a running sum of the value of each element
			# across runs (frequency)
			# square_sum will keep a running sum of the value squares (to get
			# variance later)
			# write initial values to the output
			freq_sum = dict()
			square_sum = dict()
			# store element names for event_traces file
			updated_element = simStep * [0]
			for key, element in self.__getElement.items():
				freq_sum[key] = (simStep+1) * [0]
				square_sum[key] = (simStep+1) * [0]
				if normalize:
					ele_value = element.get_current_value()
					freq_sum[key][0] = ele_value * runs
					square_sum[key][0] = ele_value*ele_value * runs
				else:
					ele_val_index = element.get_current_value_index()
					freq_sum[key][0] = ele_val_index * runs
					square_sum[key][0] = ele_val_index*ele_val_index * runs

			if simtype == 'round':
				# Store the number elements needed to be updated each round
				totalNumElements = len(self.__updateList)

			# Perform 'runs' number of simulation runs
			if progressReport:
				progress_ticks = np.linspace(0,runs,11)
			for run in range(runs):
				# Set elements to initial values
				self.set_initial(scenario)

				if simtype == 'sync' or simtype == 'sync_multi' or simtype == 'rand_sync' or simtype == 'rand_sync_guass' or randomizeEachRun == True:
					self.set_random_initial(scenario)

				# 'memo' will store each element's values for each step/round
				# TODO: a numpy array/matrix might be more efficient for memo
				memo = dict()
				memo_trend = dict()

				# write initial value to the output
				for key, element in self.__getElement.items():
					if normalize:
						memo[key] = [element.get_current_value()]
					else:
						memo[key] = [element.get_current_value_index()]
					memo_trend[key] = [0]
					element.set_trend_index(0)

					if key in self.__updateList:
						element.set_prev_value({el_reg: self.__getElement[el_reg].get_current_value() for el_reg in element.get_name_list()})
						element.set_prev_index({el_reg: self.__getElement[el_reg].get_current_value_index() for el_reg in element.get_name_list()})

					else:
						element.set_prev_value({key:element.get_current_value()})
						element.set_prev_index({key:element.get_current_value_index()})


				# Perform 'simStep' number of simulation steps (or rounds)
				for step in range(1, simStep+1):
					# Update elements according to the simulation scheme
					if simtype == 'fixed_updates':
						element = updates[run][step-1]
						name = self.fixed_updates(memo, step, element)
						#  update the element to its next state value.
						self.__getElement[name].set_value_index(
							self.__getElement[name].get_next_value_index())

						# Store element name for event_traces file
						updated_element[step-1] = name

					if simtype == 'ra':
						# randomly choose an element or element group to update, by calculating it's next state value
						name = self.ra_update(memo, step)
						if name in self.__groupUpdate:
							# calculate the next state value for all other elements in this group
							for key in self.__groupUpdate:
								if self.__groupUpdate[name] == self.__groupUpdate[key]:
									# Have to first calculate next state values of all elements
									self.update_next(key)
							for key in self.__groupUpdate:
								if self.__groupUpdate[name] == self.__groupUpdate[key]:
									# Now update the variables with next state values.
									self.__getElement[key].set_value_index(
										self.__getElement[key].get_next_value_index())
						# Finally update the originally generated random element to its next state value.
						self.__getElement[name].set_value_index(
							self.__getElement[name].get_next_value_index())

						# Store element name for event_traces file
						updated_element[step-1] = name

					elif simtype == 'sync' or simtype ==  'rand_sync' or simtype == 'rand_sync_gauss':
						# simultaneously update all elements either with random delays or not
						self.sync_update(memo, step, simtype)


					elif simtype == 'ra_multi':
						# Randomly select an element in the model. Decide whether or not to update it based on probability.
						# If the element is to be updated, also update the other elements in its group.
						if bool(self.__groupUpdate):
							# Randomly generate an element and see if it passed the probability check.
							element, passed = self.prob_update_ra()
							if element in self.__groupUpdate:
								# compute next state values for all other elements in this group
								for key in self.__groupUpdate:
									if (self.__groupUpdate[element] == self.__groupUpdate[key]):
										# only compute next state value if the first element in the group's probability check passed
										if passed:
											self.update_next(key)
										else:
											# Next state value should be the element's current state value
											self.__getElement[key].__next_val_index = self.__getElement[key].get_current_value_index()
								for key in self.__groupUpdate:
									if (self.__groupUpdate[element] == self.__groupUpdate[key]):
										# Update the current values with the computed next state values
										self.__getElement[key].set_value_index(
											self.__getElement[key].get_next_value_index())
							else:
								# If this particular element did not belong to any update groups, just update it alone with its computed next state value
								self.__getElement[element].set_value_index(
									self.__getElement[element].get_next_value_index())
						# The case if we are not using update groups at all
						else:
							element, passed = self.prob_update_ra()
							self.__getElement[element].set_value_index(
								self.__getElement[element].get_next_value_index())

					elif simtype == 'sync_multi':
						# Run through each element in the model. Decide whether or not to update the rule based on probability
						# If a set of rules belongs to a group, make the decision to update them only one time (they either all get updated or stay the same)
						# The probability for the elements belonging to a group is the same as the first element in the group.
						# First check if we are using groups (check if groupUpdate dictionary is populated)
						if bool(self.__groupUpdate):
							# Use a list to keep track of elements that have already been updated, in order to avoid redundancy
							alreadyUpdated = []
							for element in self.__updateList:
								# If this element's update group has already been computed, ignore it
								if element not in alreadyUpdated:
									passed = self.prob_update(element)
									alreadyUpdated += [element]
									if element in self.__groupUpdate:
										# compute next state values for all other elements in this group
										for key in self.__groupUpdate:
											if (self.__groupUpdate[element] == self.__groupUpdate[key]):
												# only compute next state value if the first element in the group's probability check passed
												if passed:
													self.update_next(key)
												else:
													# Next state value should be the element's current state value
													self.__getElement[key].__next_val_index = self.__getElement[key].get_current_value_index()
												# All elements in group are added to update list regardless of whether the check passed, to avoid recheck
												alreadyUpdated += [key]
							for element in self.__updateList:
								# Now update all elements based on next state values
								self.__getElement[element].set_value_index(
									self.__getElement[element].get_next_value_index())

						# The case if we are not using update groups
						else:
							for element in self.__updateList:
								# Compute all next state values for each element
								passed = self.prob_update(element)
							for element in self.__updateList:
								# Now update all elements based on next state values
								self.__getElement[element].set_value_index(
									self.__getElement[element].get_next_value_index())

					elif simtype == 'round':
						# Update all elements in the model once each round. If doing rank update simulation, update based on rank.
						# Elements can be updated as soon as their next value is calculated. Each individual element in the list
						# must be updated before a particular element can be updated again

						# Obtain the ranks in order of smallest to largest.
						# By default, the rankUpdate list will just contain 0 and all elements will have the same rank
						rankList = sorted(self.__rankUpdate.keys())
						# Place the highest rank first as that is what you want to run first
						rankList.reverse()
						# For each possible rank
						for rank in rankList:
							# Retrieve array of everything in a particular rank
							currentRankList = self.__rankUpdate[rank]
							# Create a list from 0 to num elements in rank - 1
							n = len(currentRankList)
							randomList = [i for i in range(n)]
							# Shuffle the list and use it to index and update each element of this rank in random order
							random.shuffle(randomList)
							for j in randomList:
								element = currentRankList[j]
								self.__getElement[element].update(self.__getElement, memo, step)

					# Store element values for this step
					for key, element in self.__getElement.items():

						if normalize:
							ele_value = element.get_current_value()
						else:
							ele_value = element.get_current_value_index()
						memo[key] += [ele_value]

						# Check if the element has been updated, to update the trend.
						if key == updated_element[step-1]:
							memo_trend[key] += [memo[key][-1]-memo[key][-2]]
							element.set_trend_index(memo_trend[key][-1])
						else:
							memo_trend[key] += [element.get_trend_index()]

						# Store frequency and square values
						freq_sum[key][step] += ele_value
						square_sum[key][step] += ele_value*ele_value

						# Check for element value toggles
						if key in self.__switchStep:
							for index, switch_step in enumerate(self.__switchStep[key][scenario]):
								if switch_step == step:
									# set element value to the toggle value
									toggle_val = self.__switchValue[key][scenario][index]
									# calculate new trends
									trend_current = toggle_val - element.get_current_value_index()
									element.set_trend_index(trend_current)
									element.set_value_index(toggle_val)

									if normalize:
										ele_value = element.get_current_value()
										memo[key][step] = ele_value
										freq_sum[key][step] = ele_value * runs
										square_sum[key][step] = ele_value*ele_value * runs
									else:
										ele_val_index = element.get_current_value_index()
										memo[key][step] = ele_val_index
										freq_sum[key][step] = ele_val_index * runs
										square_sum[key][step] = ele_val_index*ele_val_index * runs

								else:
									element.set_trend_index(element.get_trend_index())
								memo_trend[key][step] = element.get_trend_index()

				# Write values from this run to the output file
				if outMode == 1:
					output_file.write('Run #'+str(run)+'\n')
					for name in sorted(self.__getElement):
						if normalize:
							out_level = 2
						else:
							out_level = self.__getElement[name].get_levels()
						output_file.write(name+'|'+str(out_level)+'|'
													+ ' '+' '.join([str(x) for x in memo[name]])+'\n')

				elif outMode == 2:
					# transpose format (used by sensitivity analysis and model checking)
					if run == 0:
						output_file.write('# time ')
						output_file.write(' '.join([name for name in sorted(self.__getElement)]))
						output_file.write(' step\n')

					for step in range(simStep):
						output_file.write(str(step)+'  ')
						output_file.write(' '.join([str(memo[name][step])
													for name in sorted(self.__getElement)]))
						output_file.write(' '+str(step)+'\n')
				# write to the event_traces file
				elif outMode == 7:
					output_file.write('Run #'+str(run)+'\n')
					output_file.write(' '.join(updated_element)+'\n')

				if progressReport:
					if run in progress_ticks:
						print("Simulation",str(np.where(progress_ticks==run)[0][0]*10)+"% complete...")
			if progressReport:
				print("Simulation 100% complete.")

			if outMode != 2 and outMode != 7 and (simtype != 'sync' or simtype != 'sync_multi' or simtype != 'rand_sync' or simtype != 'rand_sync_gauss'):
				if outMode == 3:
					# Write total number of runs, to be used for plotting
					output_file.write('Run #'+str(run)+'\n')

				# Write frequency summary (sum of values for each element at each step across runs)
				output_file.write('Frequency Summary:\n')
				for name in sorted(self.__getElement):
					# also write number of levels for each element to output file so they can
					# be used to plot the traces later
					if normalize:
						out_level = 2
					else:
						out_level = self.__getElement[name].get_levels()
					output_file.write(name+'|'+str(out_level)+'|'
											+ ' '+' '.join([str(x) for x in freq_sum[name]])+'\n')

				output_file.write('\nSquares Summary:\n')
				for name in self.__getElement:
					if normalize:
						out_level = 2
					else:
						out_level = self.__getElement[name].get_levels()
					output_file.write(name+'|'+str(out_level)+'|'
											+ ' '+' '.join([str(x) for x in square_sum[name]])+'\n')

	def update(self, element):
		""" Update a specified element in the model
		"""

		if element in self.__updateList:
			self.__getElement[element].update(self.__getElement)
		else:
			logging.info('Element has no regulators')

	def update_next(self, element):
		""" Calculate the next state value of a specified element in the model
		"""

		if element in self.__updateList:
			self.__getElement[element].update_next(self.__getElement)
		else:
			logging.info('Element has no regulators')

	def ra_update(self, memo=dict(), step=0):
		""" Update a randomly-selected element
		"""
		if self.__totalPriority > 0:
			#If you are using update rate
			priorityIndex = random.randrange(self.__totalPriority)
			element = self.__rateUpdateList[priorityIndex]
			self.__getElement[element].update_next(self.__getElement, memo, step)
		else:
			# randomly choose an element to update, if not using optional column update rate
			element = random.choice(self.__updateList)
			# update the element's value (and each element's dictionary of the state of its regulators)
			# note that the "update" function below is the gateNode update function, not the Manager update
			self.__getElement[element].update_next(self.__getElement, memo, step)
		return element

	def fixed_updates(self, memo=dict(), step=0, element=''):
		""" Update an element from the event traces file
		"""
		# update the element's value (and each element's dictionary of the state of its regulators)
		# note that the "update" function below is the gateNode update function, not the Manager update
		self.__getElement[element].update_next(self.__getElement, memo, step)
		return element

	def sync_update(self, memo=dict(), step=0, simtype='sync'):
		""" Calculate the next value for each element based on the current values of its regulators.
			Then update all elements in the same step.
		"""
		for element in self.__updateList:
			self.__getElement[element].update_next(self.__getElement, memo, step, simtype)
		for element in self.__updateList:
			self.__getElement[element].set_value_index(self.__getElement[element].get_next_value_index())

	def prob_update(self, element):
		""" Get elements probability, generate a random float from [0.0,1.0), see if probability of element is greater, then update based on that
		"""
		prob = self.__probUpdate[element]
		if (prob > random.random()):
			self.__getElement[element].update_next(self.__getElement)
			return True
		return False

	def prob_update_ra(self):
		""" Randomly select an element. Get elements probability, generate a random float from [0.0,1.0), see if probability of element is greater, then update based on that.
		"""
		if self.__totalPriority > 0:
			#If you are using update rate
			priorityIndex = random.randrange(self.__totalPriority)
			element = self.__rateUpdateList[priorityIndex]
		else:
			element = random.choice(self.__updateList)
		prob = self.__probUpdate[element]
		if (prob > random.random()):
			self.__getElement[element].update_next(self.__getElement)
			return element, True
		return element, False

	## Logic expression functions
	def create_rules(self, outputFilename, scenario=0):
		""" Output logic expressions (rules) for each element
		"""

		# TODO: incorporate delays

		# Note: scenario is only needed to write initial values for the java simulator,
		# rules should be independent of initial value
		with open(outputFilename, 'w') as output_txt:

			all_names = set()
			my_string_buffer = StringIO()

			for key, element in self.__getElement.items():
				X = element.get_name()
				logging.info('Creating the rule for '+str(X))

				all_names |= set(element.get_name_list())
				element.generate_element_expression(my_string_buffer)

			# output initial values
			for name in sorted(list(all_names)):
				# convert initial value to binary
				bit_length = math.ceil(math.log(self.__getElement[name].get_levels(), 2))
				this_initial = self.__initial[name][scenario]
				this_initial_bin = '{0:b}'.format(int(this_initial)).zfill(bit_length)

				# write initial values for this element's boolean variables to the rule file
				# only use "_0", "_1", notation if more than one bit is needed to represent this element's value
				if bit_length > 1:
					for k in range(bit_length):
						# using reverse indexing to get the bits from right to left
						if int(this_initial_bin[-(k+1)]) == 0:
							output_txt.write(name+'_' + str(k) + ' = False\n')
						elif int(this_initial_bin[-(k+1)]) == 1:
							output_txt.write(name+'_' + str(k) + ' = True\n')
						else:
							raise ValueError('Error parsing initial value for '+name)
				else:
					if int(this_initial_bin[0]) == 0:
						output_txt.write(name + ' = False\n')
					elif int(this_initial_bin[0]) == 1:
						output_txt.write(name + ' = True\n')
					else:
						raise ValueError('Error parsing initial value for '+name)

			# output rules
			output_txt.write('\nRules:\n')
			output_txt.write(my_string_buffer.getvalue())

	def create_truth_tables(self, outputBaseFilename, scenario=0):
		""" Generate truth tables for each element in the model
		"""

		# TODO: incorporate delays

		for key, f in self.__getElement.items():

			output_model_file = openpyxl.Workbook()
			output_model = output_model_file.active

			regulated = f.get_name()
			# regulated element
			output_model.cell(row=1, column=1).value = regulated

			# write all regulators and their number of levels
			for index, name in enumerate(f.get_name_list()):
				output_model.cell(row=2, column=index+1).value = name
				levels = self.__getElement[name].get_levels()
				output_model.cell(row=3, column=index+1).value = levels

			max_reg_states = self.__getElement[regulated].get_levels()
			for index in range(max_reg_states):
				output_model.cell(row=4, column=index+2).value = str(index)

			input_states = f.generate_all_input_state()
			row = 5
			for state in input_states:
				output_model.cell(row=row, column=1).value = str(state)
				for index in range(max_reg_states):
					reg_val = self.__getElement[regulated].get_value_from_index(index)
					output_model.cell(row=row, column=index +2).value = str(f.evaluate_state(state + [reg_val]))
				row += 1

			output_model_file.save(filename='{}_{}.xlsx'.format(outputBaseFilename,key))

	def parse_truth_table(self, truthTable, levels):
		""" parse truth table input for the regulation function
		"""

		table = dict()

		# parse the truth table and set multi-index
		table['Regulators'] = [x for x in truthTable.iloc[0, :-2]]

		# read propagation delays
		table['Prop_delays'] = [0 for x in table['Regulators']]

		# read whether the update method is reset or no-reset
		reset = truthTable.keys()[1]
		if reset in ['reset', 'r', 'Reset', 'R']:
			table['Reset'] = 'reset'
		elif reset in ['no-reset', 'n', 'No-reset', 'no-Reset', 'N', 'no reset', 'No reset', 'no Reset']:
			table['Reset'] = 'no-reset'
		else:
			table['Reset'] = 'reset'  # reset is the default method.

		for i in range(len(table['Regulators'])-1):
			if '~' in table['Regulators'][i]:
				propagation_delay, reg_name = table['Regulators'][i].split('~')
				table['Prop_delays'][i] = propagation_delay
				table['Regulators'][i] = reg_name

		temp_table = truthTable.rename(
			columns=truthTable.iloc[0]).drop(truthTable.index[0])

		reg_delays = temp_table.iloc[0:, -1]
		reg_delays = reg_delays.replace(r'', np.nan, regex=True)
		reg_delays = reg_delays.fillna(0)

		temp_table = temp_table.drop(['regulation delays'], axis=1)

		temp_table = temp_table.set_index(list(temp_table.columns[:-1]))

		# create multidimensional numpy array from dataframe
		table_array_shape = list(map(len, temp_table.index.levels))
		table_array = np.full(table_array_shape, np.nan)
		table_array[tuple(temp_table.index.codes)] = temp_table.values.flat
		table['Table'] = table_array.astype(int)

		reg_delays_array = np.full(table_array_shape, np.nan)
		reg_delays_array[tuple(temp_table.index.codes)] = reg_delays.values.flat
		table['Reg_delays'] = reg_delays_array.astype(int)

		# TODO: check dimensions of table using levels

		return table

	def get_expression_from_truth_table(self, truthTable, A):
		logging.warn('Truth table conversion to expression not yet supported')
		# TODO: get logic expression from truth table
		# Parse into state-value mappings, then use state_to_expression
		# pseudocode below
		# how to convert from more than 2 levels?

		# levels = "regulated element levels"
		# elements =
		# input_states = [x for x in first_column_strings]

		# bit_length = int(math.ceil(math.log(levels,2)))
		# mode_to_expresion = [[] for x in range(bit_length)]

		# for state in input_states:
		# 	value = "truth table entry for this state"
		# 	for k in range(math.ceil(math.log(value+1,2))):
		# 		if value%2:
		# 			mode_to_expresion[k].append('('+self.state_to_expression(state)+')')
		# 		value = value//2

		return A

####################################################################
############ 			Element object  			 	############
####################################################################


class Element(object):
	""" Define a node object representing an element.
		Each element object has properties that include:
			activators,
			inhibitors,
			current value,
			number of activity levels,
			delay behavior,
			random delay parameters,
			optimization parameters,
			update rule parameters

	"""

	# Note: no default values are provided here to avoid conflicts with the default
	# values for a model object (Simulator)
	# TODO: use __slots__ for memory optimization


	def __init__(self,
			X, A, I,
			curr_val_index,
			levels, delays,
			balancing,
			spont_delay,
			noise,
			delta,
			opt_input,
			opt_output,
			opt_input_value,
			opt_output_value,
			opt_obj_weight,
			increment
			):

		# regulated element name
		self.__regulated = X.strip()

		# positive regulation (activation) function for this element
		# check for a truth table
		if type(A) is str:
			self.__act = re.sub('\s', '', A)
		elif type(A) is dict:
			self.__act = A['Table']
		else:
			raise ValueError('Invalid regulation function for ' + str(X))

		# negative regulation (inhibition) function for this element
		self.__inh = re.sub('\s', '', I)

		# number of discrete levels (states) for this element
		self.__levels = int(levels)

		# list of level values for this element
		self.__levels_array = np.linspace(0, 1, levels)

		# current element value
		# use the line below later if we decide to accept either index or float level
		# curr_val_index, = np.where(self.__levels_array == curr_val)
		self.__curr_val_index = int(curr_val_index)
		self.__curr_trend_index = 0

		# next element value. use in sync simulation to update all elements upon calculating next values for every element
		self.__next_val_index = 0

		# increment function for increasing/decreasing value
		self.__increment = increment

		# balancing behavior (for when regulation scores are equal)
		self.__balance = balancing
		self.__curr_balancing_delay = 0

		# delay values for spontaneous activation/inhibition for elements with no positive/negative regulators
		self.__spont = spont_delay
		self.__curr_spont_delay = 0

		# delay values for element level transition delays
		self.__delays = delays
		self.__max_delay = max(delays)
		self.__curr_delays = [0 for x in delays]

		# allowed noise
		self.__noise = noise

		# delay range in random-delay sync update scheme
		self.__delta = delta

		# creating a buffer with the same length as the max delay value
		if spont_delay != '' and balancing != '':
			max_delay = max(delays+[spont_delay]+[int(balancing[1])])
		elif spont_delay == '' and balancing != '':
			max_delay = max(delays+[int(balancing[1])])
		elif spont_delay != '' and balancing == '':
			max_delay = max(delays+[spont_delay])
		else:
			max_delay = max(delays)
		self.__reg_score = deque([],max_delay+1)

		# for sequential updates, we need to keep track of the step where the element was last updated
		self.__last_update_step = 0

		# names of this element and its regulators
		self.__name_list = list()
		self.__table_prop_delays = list()
		if type(A) is str:
			self.__name_list = self.create_name_list(X.strip(), A.strip(), I.strip())
		elif type(A) is dict:
			self.__name_list = A['Regulators']
			self.__table_prop_delays = A['Prop_delays']
			self.__table_reg_delays = A['Reg_delays']
			self.__table_curr_reg_delays = 0
			self.__old_table_indices = list()
			self.__update_method = A['Reset']
		else:
			raise ValueError('Invalid regulation function for ' + str(X))

		# dictionary mapping the names of this element and its regulators to their current values and indexes
		self.__name_to_value = dict()
		self.__name_to_index = dict()
		self.__name_to_trend = dict()
		self.__name_to_trend_index = dict()
		self.__previous_value = dict()
		self.__previous_index = dict()

		# check if the element is an input/output or not
		self.__opt_input = opt_input
		self.__opt_fixed_input = opt_input_value
		self.__opt_output = opt_output
		self.__opt_fixed_output = opt_output_value
		self.__opt_obj_weight = opt_obj_weight

	##############################################################
	# Get/set functions
	##############################################################
	# TODO: add set functions for these attributes and use to init in the Simulator

	def get_name(self):
		return self.__regulated

	def get_act(self):
		return self.__act

	def get_inh(self):
		return self.__inh

	def get_levels(self):
		return self.__levels

	def get_levels_array(self):
		return self.__levels_array

	def get_name_list(self):
		return self.__name_list

	def get_current_value_index(self):
		return self.__curr_val_index

	def get_next_value_index(self):
		return self.__next_val_index

	def get_current_value(self):
		return self.__levels_array[self.__curr_val_index]

	def get_value_from_index(self, index):
		return self.__levels_array[index]

	def get_index_from_value(self, value):
		return self.__levels_array.index(value)

	def get_delays(self):
		return self.__delays

	def get_spont(self):
		return self.__spont

	def get_balancing(self):
		return self.__balance

	def set_value_index(self, val_index):
		if val_index < self.__levels:
			self.__curr_val_index = val_index
		else:
			raise ValueError('Invalid value index for {}, '
					'must be < {} : {}'.format(self.__regulated,self.__levels,val_index))

	def set_trend_index(self, trend_index):
		if np.abs(trend_index) < self.__levels:
			self.__curr_trend_index = trend_index
		else:
			raise ValueError('Invalid trend index for {}, '
					'must be < {} : {}'.format(self.__regulated,self.__levels,trend_index))

	def set_prev_value(self, prev_dict):
		for key,val in prev_dict.items():
			self.__previous_value[key] = val

	def set_prev_index(self, prev_dict):
		for key,val in prev_dict.items():
			self.__previous_index[key] = val


	def is_input(self):
		return self.__opt_input

	def fixed_input(self):
		return self.__opt_fixed_input

	def fixed_output(self):
		return self.__opt_fixed_output

	def is_output(self):
		return self.__opt_output

	def objective_weight(self):
		return self.__opt_obj_weight

	def get_trend_index(self):
		return self.__curr_trend_index

	def get_trend_from_index(self, index):
		if index > 0:
			return self.__levels_array[index]
		else:
			return -self.__levels_array[-index]

	def get_trend(self):
		return self.get_trend_from_index(self.__curr_trend_index)

	def get_previous_value(self):
		return self.__previous_value

	def get_previous_index(self):
		return self.__previous_index


	#############################################################

	def create_name_list(self, X, A, I):
		""" Create a list of this element, activator, and inhibitor names
			TODO: after using get_model in Simulator __init__ , just use the parsed lists returned in the model
		"""

		global _VALID_CHARS

		names = set([X])

		reg_set = set()
		for reg_func in [A, I]:
			reg_list = list(re.findall('['+_VALID_CHARS+'=]+', reg_func))
			for regulator in reg_list:
				# Ignore weights and target values and parse names only
				if re.search(r'[a-zA-Z]', regulator):
					if '=' in regulator:
						reg_name, target = regulator.split('=')
					else:
						reg_name = regulator
					reg_set.add(reg_name)

		# sorting and concatenating this way so that the regulated name is always at the end
		return sorted(list(reg_set-names)) + list(names)

	def update(self, getElement, memo=dict(), step=0, simtype='sync'):
		""" Update this element's dictionary of its regulator's values,
			then update the element's value
		"""
		# TODO: simtype is sent to evaluate to control random delays, make this clearer
		# Store previous values and indices to be able to calculate trend with respect to the
		# values before update


		self.__name_to_value.clear()
		self.__name_to_index.clear()
		self.__name_to_trend.clear()
		self.__name_to_trend_index.clear()
		# update the state (values of this element and its regulators)
		for name in self.__name_list:
			# self.__name_to_trend[name] = getElement[name].get_trend()
			# self.__name_to_trend_index[name] = getElement[name].get_trend_index()
			self.__name_to_trend[name] = getElement[name].get_current_value() - self.get_previous_value()[name]
			self.__name_to_trend_index[name] = getElement[name].get_current_value_index() - self.get_previous_index()[name]
			self.__name_to_value[name] = getElement[name].get_current_value()
			self.__name_to_index[name] = getElement[name].get_current_value_index()

		self.__curr_val_index = self.evaluate(memo, step, simtype)
		self.set_prev_value({name: value for name,value in self.__name_to_value.items()})
		self.set_prev_index({name: value for name,value in self.__name_to_index.items()})

	def update_next(self, getElement, memo=dict(), step=0, simtype='sync'):
		""" Update this element's dictionary of its regulators' values,
			calculate the element's next value and store it in another variable,
			but DO NOT update the element's current value yet.
			This enables updating multiple elements in a single step,
			one must calculate all next state values based on all current state values
			before updating the elements.
		"""
		# TODO: simtype is sent to evaluate to control random delays, make this clearer
		# Store previous values and indices to be able to calculate trend with respect to the
		# values before update

		# self.__previous_value = {name: value for name,value in self.__name_to_value.items()}
		# self.__previous_index = {name: value for name,value in self.__name_to_index.items()}

		self.__name_to_value.clear()
		self.__name_to_index.clear()
		self.__name_to_trend.clear()
		self.__name_to_trend_index.clear()
		# update the state (values of this element and its regulators)
		for name in self.__name_list:
			# self.__name_to_trend[name] = getElement[name].get_trend()
			# self.__name_to_trend_index[name] = getElement[name].get_trend_index()
			self.__name_to_trend[name] = getElement[name].get_current_value() - self.get_previous_value()[name]
			self.__name_to_trend_index[name] = getElement[name].get_current_value_index() - self.get_previous_index()[name]
			self.__name_to_value[name] = getElement[name].get_current_value()
			self.__name_to_index[name] = getElement[name].get_current_value_index()


		self.__next_val_index = self.evaluate(memo, step, simtype)
		self.__previous_value = {name: value for name,value in self.__name_to_value.items()}
		self.__previous_index = {name: value for name,value in self.__name_to_index.items()}


	def evaluate(self, memo=dict(), step=0, simtype=''):
		""" Determine the value of the regulated element based on values of the inhibitors and activators.
			Uses greater/less than comparison of the regulator scores and increments/decrements the element value.
			Incorporates state-transition delays.
		"""

		# method for calcuating next-state value from regulation scores
		mapping = None

		if type(self.__act) is str:
			# parse regulation function and calculate activation score
			y_act = self.eval_reg(self.__act, 0, memo, step)
			mapping = 'increment'
		elif type(self.__act) is np.ndarray:
			# will use truth table mapping to update this element
			mapping = 'table'
		else:
			raise ValueError('Invalid regulation function for ' + str(self.__regulated))

		if type(self.__inh) is str:
			y_inh = self.eval_reg(self.__inh, 0, memo, step)
		else:
			raise ValueError('Invalid regulation function for ' + str(self.__regulated))

		# define number of levels for code readability
		max_value_index = self.__levels-1

		if mapping == 'table':
			# get the next value from the element's truth table
			X_next_index = self.eval_table(memo, step)

		elif mapping == 'increment':
			# compare the regulation scores and increment the current value

			# define values for code readability
			X_curr_index = self.__curr_val_index

			# check whether to randomize delays
			if simtype == 'rand_sync' or simtype == 'rand_sync_gauss':
				# randomized delays for spontaneous behavior
				if self.__spont != '' and self.__spont != 0:
					spont_dv = random.randint(
						int(self.__spont)-self.__delta, int(self.__spont)+self.__delta)
					if simtype == 'rand_sync_gauss':
						spont_dv = int(round(np.random.normal(
							int(self.__spont), self.__delta, 1)))
					if spont_dv < 0:
						D_spont = 0
					else:
						D_spont = spont_dv
				elif self.__spont == '':
					D_spont = ''
				elif self.__spont == 0:
					D_spont = 0

				# randomized delays for balancing behavior
				if len(self.__balance) == 2 and self.__balance[0] !=0:
					balancing = self.__balance[0]
					balancing_dv = random.randint(
						int(self.__balance[1])-self.__delta, int(self.__balance[1])+self.__delta)
					if simtype == 'rand_sync_gauss':
						balancing_dv = int(
							round(np.random.normal(int(self.__balance[1]), self.__delta, 1)))
					if balancing_dv < 0:
						D_balancing = 0
					else:
						D_balancing = balancing_dv
				else:
					balancing = ''
					D_balancing = 0

				D = list()
				for dv in self.__delays:
					if dv != 0:
						new_dv = random.randint(dv-self.__delta, dv+self.__delta)
						if simtype == 'rand_sync_gauss':
							new_dv = int(round(np.random.normal(dv, self.__delta, 1)))
						if new_dv < 0:
							D.append(0)
						else:
							D.append(new_dv)
					else:
						D.append(0)
			else:
				# spontaneous behavior
				D_spont = self.__spont
				# get balancing behavior if defined in the model, or set to defaults
				if len(self.__balance) == 2:
					balancing = self.__balance[0]
					D_balancing = int(self.__balance[1])
				else:
					balancing = ''
					D_balancing = 0
				D = self.__delays

			# if step_diff > 1, then we use sequential updates and we need to hold the old regulation scores
			if self.__last_update_step > 0:
				step_diff = step - self.__last_update_step
				if step_diff > 1 and step > 1:
					last_reg_score = self.__reg_score[-1]
					stp = 1
					while stp < step_diff:
						self.__reg_score.append(last_reg_score)
						self.__curr_delays[int(X_curr_index)] += 1
						if D_spont != '':
							self.__curr_spont_delay += 1
						if balancing != '':
							self.__curr_balancing_delay += 1
						stp += 1

			self.__last_update_step = step

			if D[int(X_curr_index)] > 0 and len(self.__reg_score) >= D[int(X_curr_index)]:
				delay_index_increase = np.linspace(
					1, D[int(X_curr_index)]+1, D[int(X_curr_index)], dtype=int, endpoint=False)
				delay_index_increase = np.negative(np.transpose(delay_index_increase))
				reg_scores_increase = [self.__reg_score[indx]
									for indx in delay_index_increase]
			else:
				reg_scores_increase = []

			if int(X_curr_index) > 0 and D[-int(X_curr_index)] > 0 and len(self.__reg_score) >= D[-int(X_curr_index)]:
				delay_index_decrease = np.linspace(
					1, D[-int(X_curr_index)]+1, D[-int(X_curr_index)], dtype=int, endpoint=False)
				delay_index_decrease = np.negative(np.transpose(delay_index_decrease))
				reg_scores_decrease = [self.__reg_score[indx]
									for indx in delay_index_decrease]
			else:
				reg_scores_decrease =[]

			if D_balancing != '' and D_balancing > 0 and len(self.__reg_score) >= D_balancing:
				delay_index_balancing = np.linspace(1, D_balancing+1, D_balancing, dtype=int, endpoint=False)
				delay_index_balancing = np.negative(np.transpose(delay_index_balancing))
				reg_scores_balancing = [self.__reg_score[indx]
									for indx in delay_index_balancing]
			else:
				reg_scores_balancing = []

			if D_spont != '' and D_spont > 0 and len(self.__reg_score) >= D_spont:
				delay_index_spont = np.linspace(1, D_spont+1, D_spont, dtype=int, endpoint=False)
				delay_index_spont = np.negative(np.transpose(delay_index_spont))
				reg_scores_spont = [self.__reg_score[indx] for indx in delay_index_spont]
			else:
				reg_scores_spont = []

			# determine next value of the regulated element,
			# based on the type of regulators and activation/inhibition scores
			# TODO: condense redundant code by rearranging if statements
			# TODO: move to a separate function

			# TODO: calculate increment from regulator scores y_act and y_inh and input degree of change
			# there is a check at the end of this function that the incremented value is within bounds [0,1]
			# note that this is set separately from the spontaneous/balancing increment, which could also be parameterized
			# default is to increment by 1
			increment = 1
			spont_increment = 1
			balance_increment = 1
			if self.__increment != 0:
				# increment linearly proportional to difference between regulation scores
				# with slope defined by increment input
				# scaling by max_value_index because scores are calculated from element values
				# which are each between 0 and 1
				if y_inh is None:
					reg_score = y_act
				elif y_act is None:
					reg_score = y_inh
				else:
					reg_score = abs(y_act - y_inh)

				increment_float = float(self.__increment)*reg_score*max_value_index
				increment = int(np.ceil(increment_float))
				spont_increment = max_value_index
				balance_increment = max_value_index

			if (self.__act) and (not self.__inh):
				# this element has only positive regulators, increase if activation > 0, or spontaneously decay
				if (y_act > 0):
					# check the state-transition delay value
					# since this is an increase, index the delays list using the current value of X
					# so if X is currently 0, and transitioning from 0 - 1, we want delays[0]
					# therefore, our index is X_curr
					score_one, = np.where(np.array(reg_scores_increase) == 1)
					# check the state transition delay value and increase
					if (len(score_one) >= (D[int(X_curr_index)]-self.__noise)) and (self.__curr_delays[int(X_curr_index)] >= D[int(X_curr_index)]):
						# increase and reset delay
						X_next_index = X_curr_index + increment
						self.__curr_delays[int(X_curr_index)] = 0
					else:
						# hold value and increment delay
						X_next_index = X_curr_index
						self.__curr_delays[int(X_curr_index)] += 1
					self.__reg_score.append(1)
				elif (y_act == 0):
					score_zero, = np.where(np.array(reg_scores_spont) == 0)
					if D_spont != '':
						# check spontaneous delay
						if (len(score_zero) >= (D_spont-self.__noise)) and (self.__curr_spont_delay >= D_spont):
							# spontaneously decay and reset spontaneous delay
							X_next_index = X_curr_index - spont_increment
							self.__curr_spont_delay = 0
						else:
							# hold value and increment spontaneous delay
							X_next_index = X_curr_index
							self.__curr_spont_delay += 1
					else:
						X_next_index = X_curr_index
					self.__reg_score.append(0)
			elif (not self.__act) and (self.__inh):
				# this element has only negative regulators, decrease if inhibition > 0, or spontaneously increase
				if (y_inh > 0):
					# check the state-transition delay value
					# since this is a decrease, negative index the delays list using the current value of X
					# So if levels=3, delays = [delay01, delay12, delay21, delay10]
					# and if X is currently 1, and transitioning from 1-0, we want delays[-1]
					# therefore, our index is -X_curr
					score_neg_one, = np.where(np.array(reg_scores_decrease) == -1)
					if (len(score_neg_one) >= (D[-int(X_curr_index)]-self.__noise)) and (self.__curr_delays[-int(X_curr_index)] >= D[-int(X_curr_index)]):
						# decrease and reset delay
						X_next_index = X_curr_index - increment
						self.__curr_delays[-int(X_curr_index)] = 0
					else:
						# hold value and increment delay
						X_next_index = X_curr_index
						self.__curr_delays[-int(X_curr_index)] += 1
					self.__reg_score.append(-1)
				elif (y_inh == 0):
					score_zero, = np.where(np.array(reg_scores_spont) == 0)
					if D_spont != '':
						# check spontaneous delay
						if (len(score_zero) >= (D_spont-self.__noise)) and (self.__curr_spont_delay >= D_spont):
							# spontaneously increase and reset spontaneous delay
							X_next_index = X_curr_index + spont_increment
							self.__curr_spont_delay = 0
						else:
							# hold value and increment spontaneous delay
							X_next_index = X_curr_index
							self.__curr_spont_delay += 1
					else:
						X_next_index = X_curr_index
					self.__reg_score.append(0)
			elif (self.__act) and (self.__inh):
				# this element has both activators and inhibitors
				# increase the value if activation > inhibition,
				# decrease if activation < inhibition,
				# check balancing if activation == inhibition
				if (y_act > y_inh):
					score_one, = np.where(np.array(reg_scores_increase) == 1)
					# check the state transition delay value and increase
					if (len(score_one) >= (D[int(X_curr_index)]-self.__noise)) and (self.__curr_delays[int(X_curr_index)] >= D[int(X_curr_index)]):
						# increase and reset delay
						X_next_index = X_curr_index + increment
						self.__curr_delays[int(X_curr_index)] = 0
					else:
						# hold value and increment delay
						X_next_index = X_curr_index
						self.__curr_delays[int(X_curr_index)] += 1
					self.__reg_score.append(1)
				elif (y_act == y_inh):
					score_zero, = np.where(np.array(reg_scores_balancing) == 0)
					# check balancing behavior since regulator scores are equal
					if balancing != '':
						if balancing in ['decrease', 'negative']:
							# check balancing delay
							if (len(score_zero) >= (D_balancing-self.__noise)) and (self.__curr_balancing_delay >= D_balancing):
								# decay and reset balancing delay
								X_next_index = X_curr_index - balance_increment
								self.__curr_balancing_delay = 0
							else:
								# hold value and increment balancing delay
								X_next_index = X_curr_index
								self.__curr_balancing_delay += 1
						elif balancing in ['increase', 'positive']:
							# check balancing delay
							if (len(score_zero) >= (D_balancing-self.__noise)) and (self.__curr_balancing_delay >= D_balancing):
								# restore and reset balancing delay
								X_next_index = X_curr_index + balance_increment
								self.__curr_balancing_delay = 0
							else:
								# hold value and increment balancing delay
								X_next_index = X_curr_index
								self.__curr_balancing_delay += 1
						else:
							raise ValueError('Invalid balancing value ' + str(balancing))
					else:
						# no balancing behavior
						X_next_index = X_curr_index
					self.__reg_score.append(0)
				elif (y_act < y_inh):
					score_neg_one, = np.where(np.array(reg_scores_decrease) == -1)
					# check the state transition delay value and decrease
					if (len(score_neg_one) >= (D[-int(X_curr_index)]-self.__noise)) and (self.__curr_delays[-int(X_curr_index)] >= D[-int(X_curr_index)]):
						# decrease and reset delay
						X_next_index = X_curr_index - increment
						self.__curr_delays[-int(X_curr_index)] = 0
					else:
						# hold value and increment delay
						X_next_index = X_curr_index
						self.__curr_delays[-int(X_curr_index)] += 1
					self.__reg_score.append(-1)
			else:
				# this element has no regulators
				X_next_index = X_curr_index
		else:
			raise ValueError('Invalid update mapping')

		# return the next state, with a check to keep X_next_index within bounds
		# if X_next_index < 0, it will return 0, as order will be X_next_index, 0, max_value_index
		# elif X_next_index > max_value_index, it will return max_value_index, as order will be 0, max_value_index, X_next_index
		# else 0 <= X_next_index <= max_value_index (within bounds), it will return X_next_index, as order will be 0, X_next_index, max_value_index
		return sorted([0, X_next_index, max_value_index])[1]

	def eval_reg(self, reg_rule, layer, memo=dict(), step=0):
		""" Calculates a regulation score based on the value of the activators or inhibitors
			(the list in reg_rule).
			Uses discrete AND, OR, and NOT (min, max, n's complement).
			Inputs:
				reg_rule : activator or inhibitor function notation
				layer : set != 0 when the function is called recursively
			Returns score y_sum
		"""
		# TODO: try creating lookup tables and index them instead of parsing every time
		# would need to incorporate delays in tables

		# Only calculate the score if there are actually regulators for this element
		if reg_rule:

			N = self.__levels-1

			y_init = list()
			y_necessary = list()
			y_enhance = list()
			y_sum = list()
			weight = 0
			summation = False

			# create a list of regulators from influence set notation
			# regulators are separated by commas or + (outside parentheses)
			# commas indicate logical OR, + indicates summation
			# Note that if 'x' in str is faster than re.search or str.find
			if '+' not in reg_rule:
				reg_list = self.split_comma_outside_parentheses(reg_rule)
			else:
				if ',' in reg_rule:
					reg_test = reg_rule.split(',')
					raise ValueError(
						'Found mixed commas (OR) and plus signs (ADD) in regulator function. Check for deprecated highest state notation element+ and replace with element^')
				elif reg_rule[-1] == '+':
					raise ValueError(
						'Check for deprecated highest state notation: replace element+ with element^')
				else:
					reg_list = reg_rule.split('+')
					# set the summation flag to indicate these should be summed
					summation = True

			# parse regulators, first checking for elements in {},  ()
			for reg_element in reg_list:

				if reg_element[0] == '{' and reg_element[-1] == '}':
					# this is an initializer
					# confirm that this is layer 0
					assert(layer == 0)
					# define as an initializer, evaluating only the expression within the brackets
					# check for a weight and multiply
					if '*' in reg_element:
						weight, name = reg_element[1:-1].split('*')
						y_init += float(weight)*self.eval_reg(name, 1, memo, step)
					else:
						y_init += self.eval_reg(reg_element[1:-1], 1, memo, step)

				elif reg_element[0] == '{' and reg_element[-1] == ']':
					# this is a necessary pair
					# Find the cut point between {} and []
					parentheses = 0
					cut_point = 0
					for index, char in enumerate(reg_element):
						if char == '{':
							parentheses += 1
						elif char == '}':
							parentheses -= 1
						if parentheses == 0:
							cut_point = index
							break

					necessary_element = reg_element[1:cut_point]
					enhance_element = reg_element[cut_point+2:-1]
					# define the first part as the necessary element
					# check for weights (asterisk notation indicating multiplication by weight)
					if '*' in necessary_element:
						weight, name = necessary_element.split('*')
						y_necessary += float(weight)*self.eval_reg(name, 1, memo, step)
					else:
						y_necessary += self.eval_reg(necessary_element, 1, memo, step)

					# define the second part as the enhancing/strengthening element
					# check for weights (asterisk notation indicating multiplication by weight)
					if '*' in enhance_element:
						weight, name = enhance_element.split('*')
						y_enhance += float(weight)*self.eval_reg(name, 1, memo, step)
					else:
						y_enhance += self.eval_reg(enhance_element, 1, memo, step)

					# increment the score according to the values of both the sufficient and enhancing elements
					# but use the 'sorted' expression to keep the value below the maximum value

					# TODO: need notation for necessary pair with sum
					# right now, uses sum if an element weight has been defined
					if float(weight) > 0:
						y_sum += [0 if all([y == 0 for y in y_necessary]) == True
													else float(sum(y_necessary)+sum(y_enhance))]
					else:
						y_sum += [0 if all([y == 0 for y in y_necessary]) == True
													else sorted([0, float(max(min(y_necessary), max(y_enhance))), 1])[1]]

				elif reg_element[0] == '(' and reg_element[-1] == ')':
					# this is a logic AND operation, all activators must be present
					# construct a list of the values of each element, then perform discrete logic AND (min)
					y_and = [float(x)
											for and_entity in self.split_comma_outside_parentheses(reg_element[1:-1])
											for x in self.eval_reg(and_entity, 1, memo, step)]
					y_sum += [min(y_and)]

				else:
					# single regulator
					# confirm that there are no commas remaining
					assert(',' not in reg_element)

					# calculate the value of the score based on the value of this regulator
					# using the element name to value mapping (__name_to_value dictionary)
					if reg_element[-1] == '^':
						# this is a highest state regulator
						# score is either zero or the maximum
						if reg_element[0] == '!':
							y_sum += [0 if self.__name_to_value[reg_element[1:-1]] == 1 else 1]
						else:
							y_sum += [1 if self.__name_to_value[reg_element[:-1]] == 1 else 0]

					elif '&' in reg_element:
						# It means that this regulation contains trend weights.
						# tag_based_eval processes single regulation string
						# with 0, 1 or more '*' (AND gates), splits factors
						# evaluates whole string and returns the result.
						# It supports trend and regular weights, max level activation,
						# target level activation, and discrete negation.
						y_sum += [self.tag_based_eval(reg_element)]

					elif '*' in reg_element:
						# parse a list of values multiplied together
						multiplied_reg_values = list()
						multiplied_reg_list = reg_element.split('*')
						for reg in multiplied_reg_list:
							if not re.search(r'[a-zA-Z]', reg):
								# this is a weight, append the value directly
								multiplied_reg_values.append(float(reg))
							else:
								# this is an element, evaluate the value and then append
								multiplied_reg_values.append(
									float(self.eval_reg(reg, 1, memo, step)[0]))
						y_sum += [np.prod(np.array(multiplied_reg_values))]
					elif reg_element[0] == '!':
						# NOT notation, uses n's complement
						if '~' in reg_element[1:]:
							propagation_delay, name = reg_element[1:].split('~')
							propagation_delay = random.randint(
								int(propagation_delay)-self.__delta, int(propagation_delay)+self.__delta)
							if int(propagation_delay) < 0:
								propagation_delay = 0
							old_values = memo[name]
							if int(propagation_delay) < len(old_values) and int(propagation_delay) != 0:
								effective_value = int(old_values[step-int(propagation_delay)])
								y_sum += [float(self.discrete_not(float(effective_value/N), 1))]
							elif int(propagation_delay) != 0:
								y_sum += [float(self.discrete_not(float(old_values[0]/N), 1))]
							else:
								y_sum += [float(self.discrete_not(float(self.__name_to_value[name]), 1))]
						else:
							y_sum += [float(self.discrete_not(float(self.__name_to_value[reg_element[1:]]), 1))]
					elif '=' in reg_element:
						# check target value
						name, target_state = reg_element.split('=')
						# add 0 or highest level to the score
						if self.__name_to_index[name] == int(target_state):
							y_sum += [1]
						else:
							y_sum += [0]
					elif '~' in reg_element:
						# check propagation delay
						propagation_delay, name = reg_element.split('~')
						propagation_delay = random.randint(
							int(propagation_delay)-self.__delta, int(propagation_delay)+self.__delta)
						if int(propagation_delay) < 0:
							propagation_delay = 0
						old_values = memo[name]
						if int(propagation_delay) < len(old_values) and int(propagation_delay) != 0:
							y_sum += [float(old_values[step-int(propagation_delay)]/N)]
						elif int(propagation_delay) != 0:
							y_sum += [float(old_values[0]/N)]
						else:
							y_sum += [float(self.__name_to_value[name])]
					else:
						y_sum += [float(self.__name_to_value[reg_element])]

			if layer == 0:
				# check for initializers and value of intializers
				if (self.__name_to_value[self.__regulated] == 0
										and len(y_init) != 0
										and all([y == 0 for y in y_init]) == True):
					return 0
				else:
					# at the top layer, sum or discrete OR (max) based on whether
					# the groups were split on commas or +
					if summation:
						return sum(y_init) + sum(y_sum)
					else:
						if len(y_sum) > 0 and len(y_init) > 0:
							return max(max(y_init), max(y_sum))
						elif len(y_sum) == 0:
							return max(y_init)
						elif len(y_init) == 0:
							return max(y_sum)
						else:
							raise ValueError('Empty y_sum and y_init')

			else:
				return y_sum

	def eval_table(self, memo=dict(), step=0):
		""" Evaluate next-state value by indexing a lookup table
		"""
		index = ()
		i = 0
		# get current values of all regulators (this includes self-regulation)
		for reg in self.__name_list:
			if int(self.__table_prop_delays[i]) > 0:
				old_values = memo[reg]
				if int(self.__table_prop_delays[i]) < len(old_values):
					self.__name_to_index[reg] = int(
						old_values[step - int(self.__table_prop_delays[i])])
				else:
					self.__name_to_index[reg] = int(old_values[0])
			# build a tuple to index np.array
			index += (self.__name_to_index[reg],)
			i += 1

		# if step_diff > 1, then we use sequential updates and we need to hold the old regulation scores
		if self.__last_update_step > 0:
			step_diff = step - self.__last_update_step
			if step_diff > 1 and step > 1 and self.__old_table_indices != []:
				last_index = self.__old_table_indices[-1]
				stp = 1
				while stp < step_diff:
					self.__old_table_indices.append(last_index)
					self.__table_curr_reg_delays += 1
					stp += 1
		self.__last_update_step = step

		# look for a regulation delay and update the index
		current_reg_delay = self.__table_reg_delays[index]
		if self.__update_method == 'reset' and current_reg_delay > 0:
			if self.__old_table_indices != []:
				if self.__old_table_indices[-1] != index:
					self.__table_curr_reg_delays = 0  # Reset delay variable
					self.__old_table_indices.clear()
			if self.__table_curr_reg_delays < current_reg_delay:
				next_state_index = self.__name_to_index[self.__name_list[-1]]
				self.__table_curr_reg_delays += 1
			else:
				next_state_index = self.__act[self.__old_table_indices[-1]]
				self.__table_curr_reg_delays = 0
				self.__old_table_indices.clear()

		elif self.__update_method == 'no-reset' and current_reg_delay > 0:
			if self.__old_table_indices != []:
				if self.__act[self.__old_table_indices[-1]] != self.__act[index]:
					self.__table_curr_reg_delays = 0  # Reset delay variable
					self.__old_table_indices.clear()
			if self.__table_curr_reg_delays < current_reg_delay:
				next_state_index = self.__name_to_index[self.__name_list[-1]]
				self.__table_curr_reg_delays += 1
			else:
				next_state_index = self.__act[index]
				self.__table_curr_reg_delays = 0
				self.__old_table_indices.clear()
		else:
			# use the values of the regulators to index the truth table
			# and get the next state value
			next_state_index = self.__act[index]
		self.__old_table_indices.append(index)
		return next_state_index

	##############################################################
	## Functions for exporting truth tables and model rules
	##############################################################
	# these functions create boolean variables for elements, create truth tables,
	# and export logic functions

	def evaluate_state(self, state):
		""" Used to determine the value of the regulated element
			for a given state (specific values of the inhibitors and activators),
			for now, this ignores delays
		"""

		# Note that much of this code is similar to evaluate(), just without delays,
		# and returns the value instead of the index

		self.__name_to_value.clear()

		for reg_index, state_val in enumerate(state):
			# create a mapping between state names and state values
			self.__name_to_value[self.__name_list[reg_index]] = float(state_val)
			self.__name_to_trend[self.__name_list[reg_index]] = 0.0

		# calculate activation and inhibition scores
		y_act = self.eval_reg(self.__act, 0)
		y_inh = self.eval_reg(self.__inh, 0)

		# define current element value, levels, and max delays for code readability
		X_curr_index = self.__levels_array.tolist().index(state[-1])
		max_value_index = self.__levels-1

		# Ignoring delays for now, but need to know spontaneous behavior
		# TODO: include delays as variables in truth tables, and
		# merge this function with evaluate()
		D_spont = self.__spont
		if len(self.__balance) == 2:
			balancing = self.__balance[0]
			D_balancing = int(self.__balance[1])
		else:
			balancing = ''
			D_balancing = 0
		D = self.__delays

		# TODO: calculate increment from regulator scores y_act and y_inh and input degree of change
		# there is a check at the end of this function that the incremented value is within bounds [0,1]
		# note that this is set separately from the spontaneous/balancing increment, which could also be parameterized
		# default is to increment by 1
		increment = 1
		if self.__increment != 0:
			# increment linearly proportional to difference between regulation scores
			# with slope defined by increment input
			# scaling by max_value_index because scores are calculated from element values
			# which are each between 0 and 1
			if y_act is not None and y_inh is None:
				reg_score = y_act*max_value_index
			elif y_act is None and y_inh is not None:
				reg_score = y_inh*max_value_index
			elif y_act is not None and y_inh is not None:
				reg_score = abs(y_act - y_inh)*max_value_index
			else:
				# no regulators
				reg_score = 0

			increment_float = float(self.__increment)*reg_score
			increment = int(np.ceil(increment_float))

		# determine next value of the regulated element,
		# based on the type of regulators and activation/inhibition scores
		if (self.__act) and (not self.__inh):
			# this element has only positive regulators, increase if activation > 0, or spontaneously decay
			if (y_act > 0):
				X_next_index = X_curr_index + increment
			elif (y_act == 0):
				if D_spont != '':
					# spontaneously decay
					X_next_index = X_curr_index - 1
				else:
					# no spontaneous behavior, hold value
					X_next_index = X_curr_index
		elif (not self.__act) and (self.__inh):
			# this element has only negative regulators, decrease if inhibition > 0, or spontaneously increase
			if (y_inh > 0):
				X_next_index = X_curr_index - increment
			elif (y_inh == 0):
				if D_spont != '':
					# spontaneously increase
					X_next_index = X_curr_index + 1
				else:
					# no spontaneous behavior, hold value
					X_next_index = X_curr_index
		elif (self.__act) and (self.__inh):
			# this element has both activators and inhibitors
			# increase the value if activation > inhibition,
			# decrease if activation < inhibition,
			# check balancing if activation == inhibition
			if (y_act > y_inh):
				X_next_index = X_curr_index + increment
			elif (y_act == y_inh):
				if balancing != '':
					if balancing in ['decrease', 'negative']:
						X_next_index = X_curr_index - 1
					elif balancing in ['increase', 'positive']:
						X_next_index = X_curr_index + 1
					else:
						raise ValueError('Invalid balancing value ' + str(balancing))
				else:
					X_next_index = X_curr_index
			elif (y_act < y_inh):
				X_next_index = X_curr_index - increment
		else:
			# this element has no regulators;
			# Note that this shouldn't happen with the current model initialization
			X_next_index = X_curr_index

		# return the next state,
		# with a sort to keep the index within bounds
		value_index = sorted([0, X_next_index, max_value_index])[1]
		value = self.__levels_array[value_index]
		return value

	def generate_all_input_state(self, include_regulated=0):
		""" Used to generate truth tables - generate and record all possible input states
			basically generating a Boolean truth table, using all possible combinations of regulator values
			the number of input states is given by levels^(number of regulators)
		"""

		# TODO: need to include delays as variables in truth tables
		# FIXME: this does not account for regulators with different numbers of levels, need to pass in getElement

		# get the total number of regulators
		# include_regulated input determines whether or not include the regulated element itself

		length = len(self.__name_list) if include_regulated else len(self.__name_list)-1
		levels = self.__levels
		total_states = []
		for num in range(int(math.pow(levels, length))):
			# generate the state
			this_state = [0]*length
			temp = num
			bit_index = -1
			while temp > 0:
				# converting the state values to the normalized range [0,1]
				this_state[bit_index] = self.__levels_array[temp % levels]
				temp = temp//levels  # integer division
				bit_index = bit_index - 1
			total_states.append(this_state)
		return total_states

	def generate_element_expression(self, output_model_file):
		""" Write logic expressions for each element in the model to an output rule file
		"""

		if self.__act == '' and self.__inh == '':
			return None

		else:
			# generate truth table
			input_states = self.generate_all_input_state(1)

			bit_length = int(math.ceil(math.log(self.__levels, 2)))
			mode_to_expression = [[] for x in range(bit_length)]

			# define the value for each state
			for state in input_states:
				value = self.evaluate_state(state)
				for k in range(math.ceil(math.log(value+1, 2))):
					if value % 2:
						mode_to_expression[k].append('('+self.state_to_expression(state)+')')
					value = value//2

			# write the model to a txt file
			output_model_file.write('{\n')

			# only use the underscore bit# notation if there is more than one bit needed for this element's value
			if bit_length > 1:
				for index in range(bit_length):
					mode = mode_to_expression[index]
					if len(mode) != 0:
						output_model_file.write(self.__regulated+'_'+str(index)+' = ' +
													' + '.join(mode)+';\n')
					else:
						output_model_file.write(
							self.__regulated+'_'+str(index)+' = Const_False;\n')
			else:
				mode = mode_to_expression[0]
				if len(mode) != 0:
					output_model_file.write(self.__regulated+' = ' +
											' + '.join(mode)+';\n')
				else:
					output_model_file.write(self.__regulated+' = Const_False;\n')

			output_model_file.write('}\n')

	def state_to_expression(self, state):
		""" Create a logical expression for the state in sum of products form
			TODO: use something like sympy to simplify the Boolean functions before writing to the output file
		"""
		result = list()

		for index, state_val in enumerate(state):
			element = self.__name_list[index]
			value = int(state_val)

			bit_length = int(math.ceil(math.log(self.__levels, 2)))

			# only use underscore bit# notation if there is more than one bit needed for this element's value
			if bit_length > 1:
				for k in range(bit_length):
					if value % 2:
						result.append(element+'_'+str(k))
					else:
						result.append('!'+element+'_'+str(k))
					value = value//2
			else:
				if value % 2:
					result.append(element)
				else:
					result.append('!'+element)

		return '*'.join(result)

	##############################################################
	## Utility functions
	##############################################################

	def discrete_not(self, x, N):
		""" Compute NOT using n's complement
		"""

		assert N >= x, 'Can\'t compute NOT, input ({}) is greater than maximum value ({})'.format(x,N)

		return (N - x)

	def split_comma_outside_parentheses(self, sentence):
		""" Parse comma-separated strings in a regulation function,
			preserving groups of elements in parentheses or brackets (AND, necessary pair, initializers).
		"""
		final_list = list()
		parentheses = 0
		start = 0
		for index, char in enumerate(sentence):
			if index == len(sentence)-1:
				final_list.append(sentence[start:index+1])
			elif char == '(' or char == '{' or char == '[':
				parentheses += 1
			elif char == ')' or char == '}' or char == ']':
				parentheses -= 1
			elif (char == ',' and parentheses == 0):
				final_list.append(sentence[start:index])
				start = index+1
		return final_list

	def tag_based_eval(self, regulation_string):
		"""Trend-based regulation evaluation function. It is tag-based to allow multiplication
		as discrete disjunction."""
		if regulation_string=='':
			return 0

		# Split the product into factors
		reg_list = regulation_string.split('*')
		# Assign tags to discriminate between element and weight factors
		tag_list = ['ELEMENT' if re.search(r'[A-Za-z\_]',elem) else 'WEIGHT'
					for elem in reg_list]

		# Initialize calculations (mul_reg stores regular weights,
		# mul_trend stores trend weights, mul_list stores all factors for multiplication)
		mul_reg = 0
		mul_trend = 0
		mul_list = [1.]
		for i, tag in enumerate(tag_list):
			if tag=='WEIGHT':
				and_count = reg_list[i].count('&')
				if and_count==1:
					trend_weight, reg_weight = reg_list[i].split('&')
					try:
						trend_weight = float(trend_weight)
						reg_weight = float(reg_weight)
					except:
						raise ValueError('Invalid weights: weights need to be integers or floats.')
					if mul_reg!=0 or mul_trend!=0:
						mul_reg *= reg_weight
						mul_trend *= trend_weight
					else:
						mul_reg = reg_weight
						mul_trend = trend_weight
				elif and_count==0:
					trend_weight = 0
					try:
						reg_weight = float(reg_list[i])
					except:
						raise ValueError('Regulator weight must be a number. Simulator encountered: '+reg_list[i])
					if mul_reg!=0:
						mul_reg *= reg_weight
					else:
						mul_reg = reg_weight
					mul_trend = 0
				else:
					raise ValueError('Too many "&" signs: there can only be one or none.')
			elif tag=='ELEMENT':
				target_lvl = -1
				if (i>0 and tag_list[i-1]=='ELEMENT') or i==0:
					mul_reg = 1
					mul_trend = 0
				if reg_list[i][0]=='!':
					reg_string = reg_list[i][1:]
					not_gate = 1
				elif reg_list[i][0]=='?':
					reg_string = reg_list[i][1:]
					not_gate = 2
				else:
					reg_string = reg_list[i]
					not_gate = 0
				if reg_list[i][-1]=='^':
					reg_string = reg_string[:-1]
					max_val = 1
				else:
					max_val = 0
				if re.search(r'=[0-9]+',reg_string):
					try:
						reg_string, target_lvl = reg_string.split('=')
						target_lvl = int(target_lvl)
					except:
						raise ValueError('Invalid target level for regulator '+str(reg_string))
				try:
					trend = float(self.__name_to_trend[reg_string])

				except:
					raise ValueError('Invalid element name '+str(reg_string))

				value = self.__name_to_value[reg_string]

				if not_gate==2:
					assert value<=1, "Invalid normalized value (>1), Boolean negation requires it to be within the levels range."
					value = int(value==0)
					trend *= -1

				if target_lvl!=-1:
					idx = self.__name_to_index[reg_string]
					value *= int(idx==target_lvl)
					trend *= int(idx==target_lvl)

				if max_val==1:
					value = int(value==1)
					trend *= int(value==1)

				if not_gate==1:
					assert value<=1, "Invalid normalized value (>1), discrete negation requires it to be within the levels range."
					value = 1 - value
					trend *= -1

				if trend*mul_trend < 0:
					mul_trend = 0
				mul_list.append(mul_reg*value+mul_trend*trend)

				mul_reg = 0
				mul_trend = 0

		return np.prod(mul_list)
